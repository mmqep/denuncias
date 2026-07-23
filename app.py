# -*- coding: utf-8 -*-
import os
import uuid
import random
import string
from datetime import datetime
from functools import wraps
from io import BytesIO

from dotenv import load_dotenv
load_dotenv()

from flask import (
    Flask,
    Blueprint,
    request,
    session,
    redirect,
    url_for,
    flash,
    render_template,
    g,
    abort,
    Response,
    current_app,
    jsonify,
)

import openpyxl
from openpyxl.utils import get_column_letter

import pytz
from pymysql.err import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from config import TIMEZONE, DB_CONFIG, ALLOWED_EXTENSIONS, UPLOAD_FOLDER, MAX_UPLOAD_SIZE_BYTES
from db import close_db, get_db

ESTADOS_PERMITIDOS = [
    "Recibida",
    "En revisión",
    "Asignada",
    "En atención",
    "Solucionada",
    "Cerrada",
    "Archivada / No procedente",
]

PRIORIDADES = ["Baja", "Media", "Alta", "Crítica"]

TZ = None


def get_tz():
    global TZ
    if TZ is None:
        TZ = pytz.timezone(TIMEZONE)
    return TZ


def now_local():
    return datetime.now(get_tz())


def obtener_subcategoria_valida(db, sid):
    try:
        sid = int(sid)
    except (TypeError, ValueError):
        return None
    q = """
    SELECT s.id, s.nombre, s.categoria_id, s.area_id_principal, s.usuario_id_principal,
           s.activo AS sub_activo,
           c.nombre AS categoria_nombre, c.activo AS cat_activo
    FROM subcategorias s
    INNER JOIN categorias c ON c.id = s.categoria_id
    WHERE s.id = %s
    LIMIT 1
    """
    with db.cursor() as cur:
        cur.execute(q, (sid,))
        return cur.fetchone()


def construir_catalogo_portal_publico(db):
    with db.cursor() as cur:
        cur.execute("SELECT id, nombre FROM categorias WHERE activo=1 ORDER BY orden, nombre")
        cats = cur.fetchall()
        out = []
        for c in cats:
            cur.execute(
                "SELECT id, nombre FROM subcategorias WHERE categoria_id=%s AND activo=1 ORDER BY orden, nombre",
                (c["id"],),
            )
            subs = cur.fetchall()
            out.append(
                {
                    "id": c["id"],
                    "nombre": c["nombre"],
                    "subs": [{"id": s["id"], "nombre": s["nombre"]} for s in subs],
                }
            )
    return out


def listar_areas_activas(db):
    """Áreas activas con datos de contacto del responsable (join usuarios)."""
    with db.cursor() as cur:
        cur.execute(
            """
            SELECT a.id, a.nombre, r.nombres AS contacto_nombre, r.correo AS contacto_correo
            FROM areas a
            LEFT JOIN usuarios r ON r.area_id = a.id AND r.rol='Responsable' AND r.activo=1
            WHERE a.activo=1
            ORDER BY a.nombre
            """
        )
        return cur.fetchall()


def resolver_usuario_responsable_area(conn, area_id):
    if not area_id:
        return None
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, nombres, correo FROM usuarios
            WHERE rol='Responsable' AND activo=1 AND area_id=%s
            LIMIT 1
            """,
            (area_id,),
        )
        return cur.fetchone()


def listar_areas_con_responsable_activo(conn, excluir_area_id=None):
    """Áreas activas con al menos un Responsable activo vinculado."""
    with conn.cursor() as cur:
        sql = """
            SELECT a.id, a.nombre, u.nombres AS contacto_nombre
            FROM areas a
            INNER JOIN usuarios u ON u.area_id = a.id AND u.rol='Responsable' AND u.activo=1
            WHERE a.activo=1
        """
        params = []
        if excluir_area_id is not None:
            sql += " AND a.id <> %s"
            params.append(int(excluir_area_id))
        sql += " ORDER BY a.nombre"
        cur.execute(sql, params)
        return cur.fetchall()


def listar_administradores_activos(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, nombres, correo FROM usuarios
            WHERE rol='Administrador' AND activo=1
            ORDER BY nombres
            """
        )
        return cur.fetchall()


def denuncia_en_estado_terminal(estado):
    return estado in ("Cerrada", "Archivada / No procedente")


def listar_responsables_catalogo(conn):
    """Responsables activos con area_id para saber si están libres u ocupados."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, nombres, correo, area_id FROM usuarios
            WHERE rol='Responsable' AND activo=1
            ORDER BY nombres
            """
        )
        rows = cur.fetchall()
    out = []
    for r in rows:
        aid = r.get("area_id")
        out.append(
            {
                "id": int(r["id"]),
                "nombres": r["nombres"],
                "correo": r["correo"],
                "area_id": int(aid) if aid is not None else None,
            }
        )
    return out


def usuario_responsable_disponible_para_area(conn, uid, area_edit_id):
    """
    area_edit_id None: nueva área — solo acepta responsable sin area_id.
    area_edit_id int: edición — sin área o ya vinculado a esta misma área.
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT area_id FROM usuarios
            WHERE id=%s AND rol='Responsable' AND activo=1
            LIMIT 1
            """,
            (uid,),
        )
        r = cur.fetchone()
    if not r:
        return False
    a = r.get("area_id")
    if a is None:
        return True
    if area_edit_id is None:
        return False
    try:
        return int(a) == int(area_edit_id)
    except (TypeError, ValueError):
        return False


def asignar_responsable_a_area(conn, area_id, usuario_id_opcional):
    """
    Un solo Responsable por área (campo usuarios.area_id).
    usuario_id_opcional vacío o None: quita responsable del área.
    """
    if not area_id:
        return False, "Área no válida."
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM areas WHERE id=%s LIMIT 1", (area_id,))
        if cur.fetchone() is None:
            return False, "Área no encontrada."

    uid = usuario_id_opcional
    try:
        uid = int(uid) if uid not in (None, "") else None
    except (TypeError, ValueError):
        uid = None

    if not uid:
        with conn.cursor() as cur:
            cur.execute("UPDATE usuarios SET area_id=NULL WHERE area_id=%s", (area_id,))
        return True, None

    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, nombres, correo, rol, activo FROM usuarios WHERE id=%s LIMIT 1",
            (uid,),
        )
        row = cur.fetchone()
    if not row or row["rol"] != "Responsable" or not row["activo"]:
        return (
            False,
            "Debe elegir un usuario activo con rol Responsable (créelo antes en «Usuarios institucionales»).",
        )

    with conn.cursor() as cur:
        cur.execute("UPDATE usuarios SET area_id=NULL WHERE area_id=%s", (area_id,))
        cur.execute("UPDATE usuarios SET area_id=NULL WHERE id=%s AND rol='Responsable'", (uid,))
        cur.execute(
            "UPDATE usuarios SET area_id=%s WHERE id=%s AND rol='Responsable'",
            (area_id, uid),
        )
    return True, None


def registrar_alerta_simple(codigo, area_id, tipo, detalle_extra=""):
    linea = "[MMQEP notificación:{}] código={} area_id={} {}".format(tipo, codigo, area_id, detalle_extra)
    try:
        current_app.logger.info(linea)
    except RuntimeError:
        pass


def contar_denuncias_para_subcategoria(conn, sid):
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS c FROM denuncias WHERE subcategoria_id = %s", (sid,))
        r = cur.fetchone()
        return int(r["c"]) if r and r.get("c") is not None else 0


def contar_denuncias_para_categoria(conn, cid):
    """Expedientes cuya categoría o subcategoría pertenece a esta categoría."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*) AS c FROM denuncias
            WHERE categoria_id = %s
               OR subcategoria_id IN (SELECT id FROM subcategorias WHERE categoria_id = %s)
            """,
            (cid, cid),
        )
        r = cur.fetchone()
        return int(r["c"]) if r and r.get("c") is not None else 0


def contar_denuncias_para_area(conn, aid):
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS c FROM denuncias WHERE area_id = %s", (aid,))
        r = cur.fetchone()
        return int(r["c"]) if r and r.get("c") is not None else 0


def contar_denuncias_abiertas_asignadas_a(conn, usuario_id):
    """Expedientes no cerrados ni archivados con usuario_asignado_id = usuario_id."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*) AS c FROM denuncias
            WHERE usuario_asignado_id = %s
              AND estado NOT IN ('Cerrada', 'Archivada / No procedente')
            """,
            (usuario_id,),
        )
        r = cur.fetchone()
        return int(r["c"]) if r and r.get("c") is not None else 0


def listar_usuarios_para_reasignacion_denuncias(conn):
    """Candidatos a recibir expedientes: Administrador o Responsable activos."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, nombres, correo, rol FROM usuarios
            WHERE activo = 1 AND rol IN ('Administrador', 'Responsable')
            ORDER BY rol DESC, nombres
            """
        )
        return cur.fetchall()


def candidato_valido_recibir_denuncias(conn, uid_candidato, uid_excluir):
    if not uid_candidato or int(uid_candidato) == int(uid_excluir):
        return False
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id FROM usuarios
            WHERE id = %s AND activo = 1 AND rol IN ('Administrador', 'Responsable')
            LIMIT 1
            """,
            (uid_candidato,),
        )
        return cur.fetchone() is not None


def reasignar_denuncias_abiertas_por_desactivacion(conn, uid_origen, uid_nuevo, admin_uid):
    """
    Pasa todos los expedientes abiertos de uid_origen a uid_nuevo y registra seguimiento.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT nombres FROM usuarios WHERE id=%s LIMIT 1", (uid_origen,))
        r0 = cur.fetchone()
        nom_orig = (r0 or {}).get("nombres") or ("ID %s" % uid_origen)
        cur.execute("SELECT nombres FROM usuarios WHERE id=%s LIMIT 1", (uid_nuevo,))
        r1 = cur.fetchone()
        nom_nuevo = (r1 or {}).get("nombres") or ("ID %s" % uid_nuevo)
        cur.execute(
            """
            SELECT id, codigo, estado, area_id FROM denuncias
            WHERE usuario_asignado_id = %s
              AND estado NOT IN ('Cerrada', 'Archivada / No procedente')
            ORDER BY id
            """,
            (uid_origen,),
        )
        rows = cur.fetchall()

    now = now_local()
    for row in rows:
        did = int(row["id"])
        est = row["estado"]
        aid = row.get("area_id")
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE denuncias
                SET usuario_asignado_id = %s, fecha_actualizacion = %s
                WHERE id = %s
                """,
                (uid_nuevo, now, did),
            )
        obs = (
            "Reasignación por desactivación del usuario «{}» (id {}). "
            "Nuevo responsable operativo del expediente: «{}» (id {}). Código expediente: {}."
        ).format(nom_orig.strip(), uid_origen, nom_nuevo.strip(), uid_nuevo, row.get("codigo") or "")
        registrar_seguimiento(
            conn,
            did,
            admin_uid,
            accion="REASIGNACION_DESACTIVACION_USUARIO",
            estado_anterior=est,
            estado_nuevo=est,
            observacion=obs[:65000],
            area_id=aid,
        )


def denuncias_por_subcategoria_map(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT subcategoria_id, COUNT(*) AS c
            FROM denuncias WHERE subcategoria_id IS NOT NULL GROUP BY subcategoria_id
            """
        )
        return {int(r["subcategoria_id"]): int(r["c"]) for r in cur.fetchall()}


def denuncias_por_area_map(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT area_id, COUNT(*) AS c
            FROM denuncias WHERE area_id IS NOT NULL GROUP BY area_id
            """
        )
        return {int(r["area_id"]): int(r["c"]) for r in cur.fetchall()}


def create_app():
    app = Flask(__name__)
    app.config.from_mapping(
        SECRET_KEY=__import__("config").SECRET_KEY,
        DB_CONFIG=DB_CONFIG,
        MAX_CONTENT_LENGTH=int(MAX_UPLOAD_SIZE_BYTES) * 5,
        UPLOAD_FOLDER=os.path.join(app.root_path, UPLOAD_FOLDER),
        JSON_ASCII=False,
    )

    # Usar DictLoader con el caché generado (solución definitiva para Vercel)
    try:
        from _templates_cache import TEMPLATES
        from jinja2 import DictLoader
        app.jinja_env.loader = DictLoader(TEMPLATES)
    except ImportError:
        # Fallback por si no existe _templates_cache.py (por ejemplo, en desarrollo local)
        from jinja2 import FileSystemLoader
        template_dir = os.path.join(app.root_path, "templates")
        app.jinja_env.loader = FileSystemLoader(template_dir)

    # ========== AGREGAR ESTAS LÍNEAS ==========
    # Cargar configuraciones de Backblaze B2
    from config import B2_ACCOUNT_ID, B2_APPLICATION_KEY, B2_BUCKET_NAME, B2_ENDPOINT_URL
    app.config['B2_ACCOUNT_ID'] = B2_ACCOUNT_ID
    app.config['B2_APPLICATION_KEY'] = B2_APPLICATION_KEY
    app.config['B2_BUCKET_NAME'] = B2_BUCKET_NAME
    app.config['B2_ENDPOINT_URL'] = B2_ENDPOINT_URL
    # ==========================================

    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    @app.teardown_appcontext
    def _close_db(exc):
        close_db(exc)

    @app.after_request
    def _charset_html_utf8(response):
        ctype = response.headers.get("Content-Type") or ""
        base = ctype.split(";")[0].strip().lower()
        if base == "text/html" and "charset=" not in ctype.lower():
            response.headers["Content-Type"] = ctype + "; charset=utf-8"
        return response

    public_bp = Blueprint("public_denuncias", __name__, url_prefix="/denuncias")
    admin_bp = Blueprint("admin_denuncias", __name__, url_prefix="/denuncias/admin")

    register_public_routes(public_bp)
    register_admin_routes(admin_bp)

    app.register_blueprint(public_bp)
    app.register_blueprint(admin_bp)

    return app


def registrar_seguimiento(
    conn,
    denuncia_id,
    usuario_id,
    accion,
    estado_anterior=None,
    estado_nuevo=None,
    observacion="",
    area_id=None,
):
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO seguimientos
              (denuncia_id, usuario_id, area_id, accion, estado_anterior, estado_nuevo, observacion, fecha_creacion)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                denuncia_id,
                usuario_id,
                area_id,
                accion[:250],
                estado_anterior[:64] if estado_anterior else None,
                estado_nuevo[:64] if estado_nuevo else None,
                observacion or None,
                now_local(),
            ),
        )


def codigo_denuncia_unico(cur):
    for _ in range(50):
        suf = "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
        codigo = "MMQ-" + suf
        cur.execute("SELECT id FROM denuncias WHERE codigo=%s LIMIT 1", (codigo,))
        if cur.fetchone() is None:
            return codigo
    raise RuntimeError("No fue posible generar código único")


def allowed_file(fname):
    if not fname or "." not in fname:
        return False
    ext = fname.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def usuario_logueado():
    return session.get("uid")


def obtener_rol():
    return session.get("rol")


def usuario_puede_ver_denuncia(d, rol, uid):
    # Administrador Global y AdminCierre ven todo
    if rol in ("AdministradorGlobal", "AdminCierre"):
        return True
    if rol == "Responsable":
        # Directamente asignado a este usuario
        if d.get("usuario_asignado_id") is not None and d.get("usuario_asignado_id") == uid:
            return True
        # Sin asignar a nadie específico, pero en el área del responsable
        ua = session.get("area_id")
        if ua and d.get("area_id") and d.get("usuario_asignado_id") is None:
            try:
                return int(d["area_id"]) == int(ua)
            except (TypeError, ValueError):
                pass
        return False
    if rol == "Supervisor":
        # Supervisor solo ve denuncias de su área
        ua = session.get("area_id")
        if ua and d.get("area_id"):
            try:
                return int(d["area_id"]) == int(ua)
            except (TypeError, ValueError):
                pass
        return False
    return False


def login_required(f):
    @wraps(f)
    def w(*args, **kwargs):
        if not usuario_logueado():
            return redirect(url_for("admin_denuncias.admin_login_form"))
        return f(*args, **kwargs)

    return w


def roles_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*args, **kwargs):
            rol_usuario = obtener_rol()
            # Si el rol requerido incluye "Administrador", permitir también "AdministradorGlobal"
            if "Administrador" in roles and rol_usuario == "AdministradorGlobal":
                return f(*args, **kwargs)
            if rol_usuario not in roles:
                flash("No tiene permiso para acceder.", "danger")
                return redirect(url_for("admin_denuncias.admin_dashboard"))
            return f(*args, **kwargs)
        return w
    return deco


def cargar_denuncia_por_id(did):
    db = get_db()
    q = """
    SELECT d.*,
           u.nombres AS asignado_nombres,
           ar.nombre AS area_nombre,
           resp.correo AS area_correo,
           resp.nombres AS area_contacto_nombre
    FROM denuncias d
    LEFT JOIN usuarios u ON u.id = d.usuario_asignado_id
    LEFT JOIN areas ar ON ar.id = d.area_id
    LEFT JOIN usuarios resp ON resp.area_id = ar.id AND resp.rol='Responsable' AND resp.activo=1
    WHERE d.id=%s
    LIMIT 1
    """
    with db.cursor() as cur:
        cur.execute(q, (did,))
        return cur.fetchone()


def register_public_routes(bp):
    @bp.route("/")
    def index():
        conn = get_db()
        catalogo = construir_catalogo_portal_publico(conn)
        if not catalogo:
            flash("No hay categorías activas configuradas para el ciudadano.", "warning")
        return render_template(
            "public/index.html",
            catalogo_cats=catalogo,
            fecha_hoy=datetime.now(get_tz()).date().isoformat(),
        )

    @bp.route("/registrar", methods=["POST"])
    def registrar():
        es_anon = request.form.get("es_anonima") == "1"
        tipo = "denuncia"
        sid_raw = request.form.get("subcategoria_id")
        ubicacion = (request.form.get("ubicacion") or "").strip()[:300]
        fecha_hecho = request.form.get("fecha_hecho") or ""
        hora_hecho = request.form.get("hora_hecho") or ""
        involucrado = (request.form.get("involucrado") or "").strip()[:300]
        descripcion = (request.form.get("descripcion") or "").strip()
        acept_ver = request.form.get("acepto_veracidad") == "1"
        acept_datos = request.form.get("acepto_datos") == "1"

        nom = iden = tel = mail = None
        if not es_anon:
            nom = (request.form.get("nombres") or "").strip()[:200] or None
            iden = (request.form.get("identificacion") or "").strip()[:40] or None
            tel = (request.form.get("telefono") or "").strip()[:40] or None
            mail = (request.form.get("correo") or "").strip()[:160] or None

        # ========== AGREGAR ESTAS LÍNEAS ==========
        latitud = request.form.get("latitud") or None
        longitud = request.form.get("longitud") or None
        if latitud and longitud:
            try:
                latitud = float(latitud)
                longitud = float(longitud)
            except ValueError:
                latitud = None
                longitud = None
        # ==========================================

        errores = []
        conn = get_db()
        sc = obtener_subcategoria_valida(conn, sid_raw)
        if not sc:
            errores.append("Seleccione una subcategoría válida.")
        elif not sc.get("sub_activo") or not sc.get("cat_activo"):
            errores.append("La categoría o subcategoría seleccionada está desactivada.")
        cat_nom = (sc["categoria_nombre"] if sc else "") or ""
        sub_nom = (sc["nombre"] if sc else "") or ""
        if len(descripcion) < 30:
            errores.append("La descripción debe tener al menos 30 caracteres.")
        if not acept_ver or not acept_datos:
            errores.append("Debe aceptar la declaración de veracidad y el tratamiento de datos.")

        fh = hh = None
        fecha_hecho = fecha_hecho.strip()
        if not fecha_hecho:
            errores.append("La fecha del hecho es obligatoria.")
        else:
            try:
                fh = datetime.strptime(fecha_hecho, "%Y-%m-%d").date()
            except ValueError:
                errores.append("Fecha del hecho no válida.")

        if hora_hecho:
            try:
                hh = datetime.strptime(hora_hecho, "%H:%M").time()
            except ValueError:
                errores.append("Hora del hecho no válida.")

        if not es_anon:
            if not nom or len(nom.strip()) < 3:
                errores.append("Nombres y apellidos son obligatorios en registros identificados.")
            elif " " not in nom.strip():
                errores.append("Indique nombres y apellidos (al menos nombre y apellido).")
            if not iden:
                errores.append("Cédula o RUC es obligatorio en registros identificados.")
            tel_limpio = (tel or "").replace(" ", "").replace("-", "")
            if not tel or len(tel_limpio) < 7:
                errores.append("Teléfono válido es obligatorio en registros identificados.")
            if not mail or "@" not in mail or "." not in mail.split("@")[-1]:
                errores.append("Correo electrónico válido es obligatorio en registros identificados.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return redirect(url_for("public_denuncias.index"))

        ua = aid_asig = None
        estado_ini = "Recibida"
        aprinc = sc.get("area_id_principal") if sc else None
        uprinc = sc.get("usuario_id_principal") if sc else None
        if uprinc:
            with conn.cursor() as _cur_up:
                _cur_up.execute(
                    "SELECT id, area_id FROM usuarios WHERE id=%s AND rol='Responsable' AND activo=1 LIMIT 1",
                    (uprinc,),
                )
                _ur = _cur_up.fetchone()
            if _ur:
                ua = _ur["id"]
                aid_asig = _ur.get("area_id")
                estado_ini = "Asignada"
            elif aprinc:
                uw = resolver_usuario_responsable_area(conn, aprinc)
                if uw:
                    ua = uw["id"]
                    aid_asig = aprinc
                    estado_ini = "Asignada"
        elif aprinc:
            uw = resolver_usuario_responsable_area(conn, aprinc)
            if uw:
                ua = uw["id"]
                aid_asig = aprinc
                estado_ini = "Asignada"

        codigo = None
        nid = None
        try:
            with conn.cursor() as cur:
                codigo = codigo_denuncia_unico(cur)
                sql = """
                INSERT INTO denuncias (
                  codigo, tipo_registro, es_anonima,
                  nombres_denunciante, identificacion_denunciante, telefono_denunciante, correo_denunciante,
                  categoria_id, subcategoria_id, categoria, subcategoria, prioridad, ubicacion,
                  fecha_hecho, hora_hecho, involucrado, descripcion,
                  estado, fecha_creacion, usuario_asignado_id, area_id,
                  latitud, longitud
                ) VALUES (
                  %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
                )
                """
                vals = (
                    codigo,
                    tipo,
                    1 if es_anon else 0,
                    nom if not es_anon else None,
                    iden if not es_anon else None,
                    tel if not es_anon else None,
                    mail if not es_anon else None,
                    sc["categoria_id"],
                    sc["id"],
                    cat_nom[:220],
                    sub_nom[:220],
                    "Media",
                    ubicacion if ubicacion else None,
                    fh,
                    hh,
                    involucrado if involucrado else None,
                    descripcion,
                    estado_ini,
                    now_local(),
                    ua,
                    aid_asig,
                    latitud,
                    longitud
                )
                cur.execute(sql, vals)
                nid = cur.lastrowid

            observ_pub = (
                "Registro recibido vía canal ciudadano."
                + (" Asignación automática por área de atención (subcategoría)." if ua else "")
            )

            registrar_seguimiento(
                conn,
                nid,
                usuario_id=None,
                accion="REGISTRO_DESDE_PORTAL_PUBLICO",
                estado_anterior=None,
                estado_nuevo=estado_ini,
                observacion=observ_pub,
                area_id=aid_asig if ua else None,
            )

            if ua and codigo:
                registrar_alerta_simple(codigo, aid_asig, "NUEVO_REGISTRO_CIUDADANO", cat_nom[:80])
            
            fs = []
            lst = request.files.getlist("evidencias")
            for fobj in lst:
                if not fobj or not fobj.filename:
                    continue
                fs.append(fobj)

            for fobj in fs:
                fname = secure_filename(os.path.basename(fobj.filename))
                if not fname or not allowed_file(fname):
                    continue
                
                # Validar tamaño
                fobj.seek(0, os.SEEK_END)
                tam = fobj.tell()
                if tam > MAX_UPLOAD_SIZE_BYTES:
                    flash("Un archivo excedió el tamaño máximo permitido y fue omitido.", "warning")
                    continue
                fobj.seek(0)
                
                mime = getattr(fobj, "mimetype", None) or ""
                
                # Subir a Backblaze B2
                from cloud_storage import subir_archivo_b2

                archivo_url, nombre_guardado = subir_archivo_b2(fobj, fname)
                
                if archivo_url:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO archivos_adjuntos (
                              denuncia_id, nombre_original, nombre_guardado, ruta, tipo_mime, tamano, fecha_creacion
                            ) VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (nid, fname, nombre_guardado, archivo_url, mime, tam, now_local()),
                        )

                else:
                    flash("No se pudo guardar un archivo en la nube.", "danger")
                    continue

            conn.commit()

        except Exception:
            conn.rollback()
            flash("No se pudo registrar la solicitud en este momento.", "danger")
            return redirect(url_for("public_denuncias.index"))

        # Notificación por correo al Responsable asignado (nunca bloquea el registro)
        try:
            from mail_service import enviar_notificacion_nueva_denuncia
            import logging as _log
            _log_mail = _log.getLogger("mmqep.mail")
            _log_mail.debug(
                "Verificación notificación: ua=%s, codigo=%s, nid=%s",
                ua, codigo, nid,
            )
            if ua and codigo and nid:
                with conn.cursor() as _cur_resp:
                    _cur_resp.execute(
                        "SELECT nombres, correo FROM usuarios WHERE id=%s LIMIT 1",
                        (ua,),
                    )
                    _resp = _cur_resp.fetchone()
                _log_mail.debug(
                    "Responsable id=%s: encontrado=%s, tiene_correo=%s",
                    ua, _resp is not None,
                    bool(_resp.get("correo")) if _resp else False,
                )
                if _resp and _resp.get("correo"):
                    _area_nombre = None
                    if aid_asig:
                        with conn.cursor() as _cur_ar:
                            _cur_ar.execute(
                                "SELECT nombre FROM areas WHERE id=%s LIMIT 1",
                                (aid_asig,),
                            )
                            _ar = _cur_ar.fetchone()
                            _area_nombre = _ar["nombre"] if _ar else None
                    _ok = enviar_notificacion_nueva_denuncia(
                        _resp["correo"],
                        _resp["nombres"],
                        {
                            "codigo": codigo,
                            "categoria": cat_nom,
                            "subcategoria": sub_nom,
                            "fecha_creacion": now_local(),
                            "estado": estado_ini,
                            "prioridad": "Media",
                            "area_nombre": _area_nombre or "",
                            "ubicacion": ubicacion if ubicacion else None,
                            "es_anonima": es_anon,
                            "nombres_denunciante": nom if not es_anon else None,
                        },
                    )
                    _log_mail.debug("Resultado envío correo: %s", "exitoso" if _ok else "fallido")
                    _obs_correo = (
                        "El sistema notificó mediante correo electrónico al Responsable asignado."
                        if _ok else
                        "No fue posible enviar la notificación por correo electrónico al Responsable."
                    )
                    registrar_seguimiento(
                        conn, nid, usuario_id=None, accion="NOTIFICACION_CORREO",
                        estado_anterior=estado_ini, estado_nuevo=estado_ini,
                        observacion=_obs_correo, area_id=aid_asig,
                    )
                    conn.commit()
                else:
                    _log_mail.warning(
                        "Responsable (id=%s) sin correo registrado; sin notificación para denuncia %s.",
                        ua, codigo,
                    )
            else:
                _log_mail.debug(
                    "Notificación omitida (denuncia no asignada automáticamente): "
                    "ua=%s, codigo=%s, nid=%s", ua, codigo, nid,
                )
        except Exception:
            import logging as _log, traceback as _tb
            _log.getLogger("mmqep.mail").error(
                "Error en bloque de notificación para denuncia %s:\n%s",
                codigo, _tb.format_exc(),
            )

        return redirect(url_for("public_denuncias.resultado_codigo", cod=codigo))

    @bp.route("/resultado")
    def resultado_codigo():
        cod = (request.args.get("cod") or "").strip().upper()
        if not cod or len(cod) > 48:
            return redirect(url_for("public_denuncias.consulta"))
        return render_template("public/resultado.html", codigo=cod, titulo_confirmacion=True)

    @bp.route("/registro/<path:cod>/comprobante.pdf")
    def comprobante_pdf(cod):
        """Comprobante con los datos guardados; el código actúa como referencia de acceso."""
        cod = (cod or "").strip().upper()
        if not cod or len(cod) > 48 or not cod.startswith("MMQ-"):
            abort(404)
        db = get_db()
        row = None
        adj = []
        with db.cursor() as cur:
            cur.execute(
                """
                SELECT codigo, tipo_registro, es_anonima, nombres_denunciante, identificacion_denunciante,
                  telefono_denunciante, correo_denunciante, categoria, subcategoria, prioridad, ubicacion,
                  fecha_hecho, hora_hecho, involucrado, descripcion, estado, fecha_creacion
                FROM denuncias WHERE codigo=%s LIMIT 1
                """,
                (cod,),
            )
            row = cur.fetchone()
            if row:
                cur.execute(
                    """
                    SELECT a.nombre_original FROM archivos_adjuntos a
                    INNER JOIN denuncias d ON d.id = a.denuncia_id
                    WHERE d.codigo=%s ORDER BY a.id ASC
                    """,
                    (cod,),
                )
                for r in cur.fetchall() or []:
                    if r.get("nombre_original"):
                        adj.append(r["nombre_original"])
        if not row:
            abort(404)
        try:
            from pdf_comprobante import construir_pdf_comprobante_denuncia

            pdf_bytes = construir_pdf_comprobante_denuncia(row, adj)
        except ImportError:
            flash(
                "El comprobante PDF no está disponible en este servidor hasta instalar: "
                "pip install reportlab Pillow (dentro del entorno virtual).",
                "warning",
            )
            return redirect(url_for("public_denuncias.resultado_codigo", cod=cod))
        except Exception:
            current_app.logger.exception("comprobante_pdf")
            flash("No se pudo generar el PDF en este momento. Intente más tarde o use la consulta por código.", "danger")
            return redirect(url_for("public_denuncias.resultado_codigo", cod=cod))
        fname = "MMQEP_comprobante_{}.pdf".format(cod.replace("/", "-"))
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="{}"'.format(fname)},
        )

    @bp.route("/consulta", methods=["GET", "POST"])
    def consulta():
        if request.method == "GET":
            return render_template("public/consulta.html")

        cod = (request.form.get("codigo") or "").strip().upper()
        errores = []
        if len(cod) < 8:
            errores.append("Ingrese un código válido.")
        db = get_db()
        den = None
        if not errores:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT codigo, estado, fecha_creacion, fecha_cierre, fecha_actualizacion, es_anonima FROM denuncias WHERE codigo=%s AND activo=1 LIMIT 1",
                    (cod,),
                )
                den = cur.fetchone()
            if den is None:
                errores.append("No existe una denuncia con ese código.")
        return render_template("public/consulta.html", errores=errores, den=den, codigo_usado=cod)


def register_admin_routes(bp):

    def listar_where_base(rol, uid):
        base_filtro = "AND d.activo = 1"
        # Administrador Global y AdminCierre ven todo
        if rol in ("AdministradorGlobal", "AdminCierre"):
            return base_filtro, []
        ua = session.get("area_id")
        if rol == "Responsable" and ua:
            return (
                base_filtro + " AND (d.usuario_asignado_id = %s OR (d.area_id = %s AND d.usuario_asignado_id IS NULL))",
                [uid, ua],
            )
        if rol == "Responsable":
            return base_filtro + " AND d.usuario_asignado_id = %s", [uid]
        if rol == "Supervisor" and ua:
            return base_filtro + " AND d.area_id IS NOT NULL AND d.area_id = %s", [ua]
        return "AND 1=0", []

    def sql_y_args_bandeja_denuncias(req_args, rol, uid, limit=800):
        """Misma lógica de filtros que la bandeja; devuelve (sql, args) para listar denuncias."""
        filt, args = listar_where_base(rol, uid)
        args = list(args)
        filtros_dyn = []

        fc = req_args.get("filtro_codigo")
        es = req_args.get("filtro_estado")
        pr = req_args.get("filtro_prioridad")
        resp = req_args.get("filtro_responsable")
        fd = req_args.get("fecha_desde")
        fh = req_args.get("fecha_hasta")

        if fc:
            filtros_dyn.append("AND d.codigo LIKE %s")
            args.append("%{}%".format(fc.strip()))

        try:
            fcid = req_args.get("filtro_categoria_id")
            if fcid:
                filtros_dyn.append("AND d.categoria_id = %s")
                args.append(int(fcid))
        except (ValueError, TypeError):
            pass

        try:
            fsid = req_args.get("filtro_subcategoria_id")
            if fsid:
                filtros_dyn.append("AND d.subcategoria_id = %s")
                args.append(int(fsid))
        except (ValueError, TypeError):
            pass

        try:
            faid = req_args.get("filtro_area_id")
            if faid:
                filtros_dyn.append("AND d.area_id = %s")
                args.append(int(faid))
        except (ValueError, TypeError):
            pass

        if es and es in ESTADOS_PERMITIDOS:
            filtros_dyn.append("AND d.estado = %s")
            args.append(es)

        if pr and pr in PRIORIDADES:
            filtros_dyn.append("AND d.prioridad = %s")
            args.append(pr)

        try:
            if resp and rol in ("Administrador", "Consulta"):
                filtros_dyn.append("AND d.usuario_asignado_id = %s")
                args.append(int(resp))
        except (ValueError, TypeError):
            pass

        if fd:
            try:
                fecha_d = datetime.strptime(fd, "%Y-%m-%d").date()
                filtros_dyn.append("AND DATE(d.fecha_creacion) >= %s")
                args.append(fecha_d)
            except ValueError:
                pass

        if fh:
            try:
                fecha_h = datetime.strptime(fh, "%Y-%m-%d").date()
                filtros_dyn.append("AND DATE(d.fecha_creacion) <= %s")
                args.append(fecha_h)
            except ValueError:
                pass

        extra = "".join(filtros_dyn)
        lim = int(limit)
        if lim < 1:
            lim = 800
        if lim > 1200:
            lim = 1200

        lista_sql = """
        SELECT d.id, d.codigo, d.estado, d.prioridad, d.es_anonima, d.fecha_creacion,
            d.subcategoria, d.categoria, u.nombres AS asignado,
            COALESCE(ar.nombre, '') AS area_nombre,
            d.ubicacion, d.descripcion, d.activo
        FROM denuncias d
        LEFT JOIN usuarios u ON u.id = d.usuario_asignado_id
        LEFT JOIN areas ar ON ar.id = d.area_id
        WHERE 1=1 {filt}{extra}
        ORDER BY d.fecha_creacion DESC
        LIMIT {lim}
        """.format(
            filt=filt,
            extra=extra,
            lim=lim,
        )
        return lista_sql, tuple(args)

    @bp.route("/login", methods=["GET"])
    def admin_login_form():
        if usuario_logueado():
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        return render_template("auth/login.html")

    @bp.route("/login", methods=["POST"])
    def admin_login_post():
        corr = (request.form.get("correo") or "").strip().lower()
        pwd = request.form.get("password") or ""
        db = get_db()
        with db.cursor() as cur:
            cur.execute(
                "SELECT id, nombres, correo, password_hash, rol, activo, area_id FROM usuarios WHERE correo=%s LIMIT 1",
                (corr,),
            )
            row = cur.fetchone()
        if not row or not row["activo"] or not check_password_hash(row["password_hash"], pwd):
            flash("Credenciales incorrectas.", "danger")
            return redirect(url_for("admin_denuncias.admin_login_form"))
        session["uid"] = row["id"]
        session["nombres"] = row["nombres"]
        session["correo"] = row["correo"]
        session["rol"] = row["rol"]
        session["area_id"] = row.get("area_id")
        flash("Bienvenido/a.", "success")
        return redirect(url_for("admin_denuncias.admin_dashboard"))

    @bp.route("/logout")
    def admin_logout():
        session.clear()
        flash("Sesión cerrada.", "info")
        return redirect(url_for("admin_denuncias.admin_login_form"))

    @bp.route("/dashboard")
    @login_required
    def admin_dashboard():
        db = get_db()
        uid = session["uid"]
        rol = session["rol"]

        filt, args = listar_where_base(rol, uid)

        estado_counts = []
        sql = """
        SELECT estado, COUNT(*) AS c
        FROM denuncias d
        WHERE 1=1 {}
        GROUP BY estado
        ORDER BY estado
        """.format(
            filt
        )

        sql_rec = """
        SELECT id, codigo, categoria, subcategoria, estado, prioridad, fecha_creacion
        FROM denuncias d
        WHERE 1=1 {}
        ORDER BY fecha_creacion DESC
        LIMIT 8
        """.format(filt)

        with db.cursor() as cur:
            cur.execute(sql, tuple(args))
            estado_counts = cur.fetchall()

            cur.execute(sql_rec, tuple(args))
            recientes = cur.fetchall()

        return render_template(
            "admin/dashboard.html",
            estado_counts=estado_counts,
            recientes=recientes,
        )

    @bp.route("/denuncias")
    @login_required
    def admin_lista_denuncias():
        db = get_db()
        rol = session["rol"]
        uid = session["uid"]

        lista_sql, args = sql_y_args_bandeja_denuncias(request.args, rol, uid, 800)

        usuarios_resp = []
        areas_opciones = []
        cats_filtro = []
        subs_filtro_raw = []

        if rol in ("Administrador", "Consulta"):
            with db.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, nombres FROM usuarios
                    WHERE rol IN ('Administrador','Responsable') AND activo=1
                    ORDER BY nombres
                    """
                )
                usuarios_resp = cur.fetchall()
                areas_opciones = listar_areas_activas(db)
                cur.execute("SELECT id, nombre FROM categorias WHERE activo=1 ORDER BY orden, nombre")
                cats_filtro = cur.fetchall()
                cur.execute(
                    """
                    SELECT s.id, s.nombre, s.categoria_id
                    FROM subcategorias s
                    INNER JOIN categorias c ON c.id = s.categoria_id
                    WHERE s.activo = 1 AND c.activo = 1
                    ORDER BY s.categoria_id, s.orden, s.nombre
                    """
                )
                subs_filtro_raw = cur.fetchall()

        with db.cursor() as cur:
            cur.execute(lista_sql, tuple(args))
            denuncias = cur.fetchall()

        subs_por_cat = {}
        for s in subs_filtro_raw:
            subs_por_cat.setdefault(s["categoria_id"], []).append({"id": s["id"], "nombre": s["nombre"]})

        return render_template(
            "admin/denuncias.html",
            denuncias=denuncias,
            estados_lista=ESTADOS_PERMITIDOS,
            prioridades_lista=PRIORIDADES,
            categorias_dropdown=cats_filtro,
            subs_por_cat=subs_por_cat,
            areas_opciones=areas_opciones,
            filtros=request.args.to_dict(),
            usuarios_resp=usuarios_resp,
        )

    def _admin_pdf_denuncia_core(did, es_externo):
        from pdf_reporte_bandeja import construir_pdf_reporte_denuncia_individual

        d = cargar_denuncia_por_id(did)
        if not d:
            abort(404)
        if not usuario_puede_ver_denuncia(d, obtener_rol(), session["uid"]):
            abort(403)

        conn = get_db()
        sql_seg = (
            """
            SELECT s.denuncia_id, s.fecha_creacion, s.accion, s.estado_anterior, s.estado_nuevo,
                   s.observacion, u.nombres AS usuario_nombre, a.nombre AS area_nombre
            FROM seguimientos s
            LEFT JOIN usuarios u ON u.id = s.usuario_id
            LEFT JOIN areas a ON a.id = s.area_id
            WHERE s.denuncia_id = %s
            """
        )
        params = [did]
        if es_externo:
            sql_seg += " AND s.accion = %s"
            params.append("GESTION_REALIZADA")
        sql_seg += " ORDER BY s.fecha_creacion ASC, s.id ASC"
        with conn.cursor() as cur:
            cur.execute(sql_seg, tuple(params))
            rows = cur.fetchall()

        dr = dict(d)
        if dr.get("asignado") is None:
            dr["asignado"] = dr.get("asignado_nombres")

        gen = session.get("nombres") or session.get("correo") or "Usuario"
        fgen = now_local().strftime("%d/%m/%Y %H:%M")
        cod = (d.get("codigo") or str(did)).replace("/", "-")
        try:
            pdf_bytes = construir_pdf_reporte_denuncia_individual(
                dr,
                rows,
                modo_externo=es_externo,
                generado_por=gen,
                fecha_emision=fgen,
            )
        except ImportError:
            flash("Instale las dependencias del PDF: pip install reportlab Pillow", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        except Exception:
            current_app.logger.exception("reporte_denuncia_pdf")
            flash("No se pudo generar el PDF en este momento.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        stamp = now_local().strftime("%Y%m%d_%H%M")
        suf = "externo" if es_externo else "interno"
        fname = "MMQEP_{}_{}_{}.pdf".format(cod, suf, stamp)
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="{}"'.format(fname)},
        )

    @bp.route("/denuncias/<int:did>/reporte-interno.pdf")
    @login_required
    @roles_required("Administrador")
    def admin_reporte_denuncia_interno_pdf(did):
        return _admin_pdf_denuncia_core(did, False)

    @bp.route("/denuncias/<int:did>/reporte-externo.pdf")
    @login_required
    @roles_required("Administrador")
    def admin_reporte_denuncia_externo_pdf(did):
        return _admin_pdf_denuncia_core(did, True)

    @bp.route("/denuncias/<int:did>")
    @login_required
    def admin_detalle(did):
        d = cargar_denuncia_por_id(did)
        if not d:
            abort(404)
        rol = obtener_rol()
        uid = session["uid"]

        if not usuario_puede_ver_denuncia(d, rol, uid):
            abort(403)

        db = get_db()
        hist = []

        usuarios_internos = []
        with db.cursor() as cur:
            cur.execute(
                """
                SELECT s.*, u.nombres AS usuario_evento_nombre, a.nombre AS area_evento_nombre
                FROM seguimientos s
                LEFT JOIN usuarios u ON u.id = s.usuario_id
                LEFT JOIN areas a ON a.id = s.area_id
                WHERE s.denuncia_id=%s
                ORDER BY s.fecha_creacion DESC, s.id DESC
                """,
                (did,),
            )
            hist = cur.fetchall()

            cur.execute(
                """
                SELECT nombre_original, nombre_guardado, tipo_mime, tamano FROM archivos_adjuntos
                WHERE denuncia_id=%s ORDER BY fecha_creacion DESC
                """,
                (did,),
            )
            archivos = cur.fetchall()

            from cloud_storage import obtener_url_archivo_b2
            for a in archivos:
                a['url_temporal'] = obtener_url_archivo_b2(a['nombre_guardado'], expiracion_segundos=86400)

            if rol == "AdministradorGlobal":

                cur.execute(
                    """
                    SELECT a.id, a.nombre,
                           u.nombres AS contacto_nombre,
                           u.correo AS contacto_correo
                    FROM areas a
                    LEFT JOIN usuarios u ON u.area_id = a.id AND u.rol='Responsable' AND u.activo=1
                    WHERE a.activo=1
                    ORDER BY a.nombre
                    """
                )
                usuarios_internos = cur.fetchall()

        reasign_areas = []
        reasign_admins = []
        if rol == "Responsable":
            ex_area = d.get("area_id")
            ex_area = int(ex_area) if ex_area is not None else None
            reasign_areas = listar_areas_con_responsable_activo(db, ex_area)
            reasign_admins = listar_administradores_activos(db)

        return render_template(
            "admin/detalle_denuncia.html",
            d=d,
            historial=hist,
            archivos=archivos,
            estados_permitidos=ESTADOS_PERMITIDOS,
            areas_asignacion=usuarios_internos,
            rol=rol,
            reasign_areas=reasign_areas,
            reasign_admins=reasign_admins,
        )

    @bp.route("/denuncias/<int:did>/asignar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_asignar(did):
        errores = []
        try:
            aid = int(request.form.get("area_id") or "")
        except ValueError:
            aid = None

        conn = get_db()
        doc = cargar_denuncia_por_id(did)
        if not doc:
            abort(404)

        if aid is None:
            flash("Seleccione un área institucional.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        with conn.cursor() as cur:
            cur.execute("SELECT id FROM areas WHERE id=%s AND activo=1", (aid,))
            if cur.fetchone() is None:
                errores.append("Área no válida.")

        urow = resolver_usuario_responsable_area(conn, aid)
        if urow is None:
            errores.append(
                "Ese área no tiene un usuario con rol Responsable activo. Cree o vincule la cuenta en «Usuarios»."
            )

        estado_ant = doc["estado"]
        estado_sig = estado_ant if estado_ant in ("Recibida", "En revisión", "Asignada", "En atención") else estado_ant

        if estado_ant in ("Recibida", "En revisión"):
            estado_sig = "Asignada"

        if errores:
            flash(errores[0], "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        nu = urow["id"]

        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE denuncias
                    SET usuario_asignado_id=%s, area_id=%s, estado=%s, fecha_actualizacion=%s
                    WHERE id=%s
                    """,
                    (nu, aid, estado_sig, now_local(), did),
                )
                registrar_seguimiento(
                    conn,
                    did,
                    session["uid"],
                    accion="ASIGNACION_AREA",
                    estado_anterior=estado_ant,
                    estado_nuevo=estado_sig,
                    observacion="Caso asignado al área institucional y a su usuario Responsable.",
                    area_id=aid,
                )

            conn.commit()
            flash("Asignación por área actualizada.", "success")
            registrar_alerta_simple(doc["codigo"], aid, "REASIGNACION_AREA", doc.get("categoria", ""))
        except Exception:
            conn.rollback()
            flash("Error al asignar.", "danger")

        return redirect(url_for("admin_denuncias.admin_detalle", did=did))
    
    @bp.route("/denuncias/<int:did>/reasignar-responsable", methods=["POST"])
    @login_required
    def admin_reasignar_por_responsable(did):
        """Responsable: pasa el expediente a otra área (su responsable) o a un administrador."""
        rol = obtener_rol()
        if rol != "Responsable":
            flash("Solo el perfil Responsable puede usar esta reasignación.", "danger")
            return redirect(url_for("admin_denuncias.admin_lista_denuncias"))

        conn = get_db()
        doc = cargar_denuncia_por_id(did)
        if not doc:
            abort(404)
        if not usuario_puede_ver_denuncia(doc, rol, session["uid"]):
            abort(403)

        if denuncia_en_estado_terminal(doc["estado"]):
            flash("No puede reasignar un expediente cerrado o archivado.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        comentario = (request.form.get("comentario_reasignacion_interno") or "").strip()
        if len(comentario) < 10:
            flash(
                "El comentario de uso interno es obligatorio (al menos 10 caracteres).",
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        tipo = (request.form.get("tipo_destino_reasignacion") or "").strip()
        if tipo not in ("area", "administrador"):
            flash("Indique si reasigna a otra área o a un administrador.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        uid_act = int(session["uid"])
        area_ref_obs = session.get("area_id") or doc.get("area_id")
        estado_ant = doc["estado"]
        
        # Guardar el responsable anterior antes de cambiarlo
        responsable_anterior = doc.get("usuario_asignado_id")

        try:
            if tipo == "area":
                try:
                    aid = int(request.form.get("area_destino_id") or "")
                except ValueError:
                    aid = None
                if aid is None:
                    flash("Seleccione el área de destino.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                doc_area = doc.get("area_id")
                if doc_area is not None and int(aid) == int(doc_area):
                    flash("Debe elegir un área distinta a la que tiene el expediente actualmente.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                valid_ids = {
                    int(r["id"])
                    for r in listar_areas_con_responsable_activo(
                        conn, int(doc_area) if doc_area is not None else None
                    )
                }
                if aid not in valid_ids:
                    flash("El área indicada no es válida o no tiene responsable activo.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                urow = resolver_usuario_responsable_area(conn, aid)
                if not urow:
                    flash("No se encontró usuario Responsable para el área elegida.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                estado_sig = estado_ant
                if estado_ant in ("Recibida", "En revisión"):
                    estado_sig = "Asignada"

                nu = int(urow["id"])
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        UPDATE denuncias
                        SET usuario_asignado_id=%s, area_id=%s, estado=%s, fecha_actualizacion=%s, 
                            responsable_anterior_id=%s, reasignado_por_id=%s, fecha_reasignacion=%s
                        WHERE id=%s
                        """,
                        (nu, aid, estado_sig, now_local(), responsable_anterior, uid_act, now_local(), did),
                    )

                registrar_seguimiento(
                    conn,
                    did,
                    uid_act,
                    accion="OBSERVACION_INTERNA",
                    estado_anterior=estado_ant,
                    estado_nuevo=None,
                    observacion=comentario,
                    area_id=area_ref_obs,
                )
                with conn.cursor() as cur:
                    cur.execute("SELECT nombre FROM areas WHERE id=%s LIMIT 1", (aid,))
                    an = cur.fetchone()
                nom_area = (an or {}).get("nombre") or ("id %s" % aid)
                obs_r = (
                    "Reasignación de expediente (perfil Responsable): queda a cargo del área «{}» "
                    "(id {}). Responsable operativo asignado: «{}» (id {}). Responsable anterior: «{}». Código: {}."
                ).format(
                    nom_area.strip(),
                    aid,
                    (urow.get("nombres") or "").strip(),
                    nu,
                    (doc.get("asignado_nombres") or "Sin responsable"),
                    doc.get("codigo") or "",
                )
                registrar_seguimiento(
                    conn,
                    did,
                    uid_act,
                    accion="REASIGNACION_POR_RESPONSABLE",
                    estado_anterior=estado_ant,
                    estado_nuevo=estado_sig,
                    observacion=obs_r[:65000],
                    area_id=aid,
                )
                conn.commit()
                registrar_alerta_simple(doc["codigo"], aid, "REASIGNACION_AREA", doc.get("categoria", "") or "")
                flash("Expediente reasignado al área indicada. Quedó registro en el historial.", "success")
                return redirect(url_for("admin_denuncias.admin_lista_denuncias"))
            else:
                try:
                    adm_id = int(request.form.get("admin_destino_id") or "")
                except ValueError:
                    adm_id = None
                if adm_id is None:
                    flash("Seleccione el administrador de destino.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, nombres FROM usuarios
                        WHERE id=%s AND rol='Administrador' AND activo=1
                        LIMIT 1
                        """,
                        (adm_id,),
                    )
                    adm = cur.fetchone()
                if not adm:
                    flash("El usuario elegido no es un administrador activo.", "danger")
                    return redirect(url_for("admin_denuncias.admin_detalle", did=did))

                estado_sig = estado_ant
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        UPDATE denuncias
                        SET usuario_asignado_id=%s, area_id=NULL, estado=%s, fecha_actualizacion=%s, 
                            responsable_anterior_id=%s, reasignado_por_id=%s, fecha_reasignacion=%s
                        WHERE id=%s
                        """,
                        (adm_id, estado_sig, now_local(), responsable_anterior, uid_act, now_local(), did),
                    )

                registrar_seguimiento(
                    conn,
                    did,
                    uid_act,
                    accion="OBSERVACION_INTERNA",
                    estado_anterior=estado_ant,
                    estado_nuevo=None,
                    observacion=comentario,
                    area_id=area_ref_obs,
                )
                nom_adm = (adm.get("nombres") or "").strip()
                obs_r = (
                    "Reasignación de expediente (perfil Responsable): queda a cargo del administrador "
                    "«{}» (id {}). Responsable anterior: «{}». Código: {}."
                ).format(nom_adm, adm_id, (doc.get("asignado_nombres") or "Sin responsable"), doc.get("codigo") or "")
                registrar_seguimiento(
                    conn,
                    did,
                    uid_act,
                    accion="REASIGNACION_POR_RESPONSABLE",
                    estado_anterior=estado_ant,
                    estado_nuevo=estado_sig,
                    observacion=obs_r[:65000],
                    area_id=None,
                )
                conn.commit()
                flash(
                    "Expediente reasignado al administrador indicado. Quedó registro en el historial.",
                    "success",
                )
                return redirect(url_for("admin_denuncias.admin_lista_denuncias"))

        except Exception as e:
            conn.rollback()
            flash("Error al reasignar el expediente.", "danger")

        return redirect(url_for("admin_denuncias.admin_detalle", did=did))

    def _usuario_puede_cambiar_estado(rol, estado_anterior, estado_nuevo):
        if estado_nuevo not in ESTADOS_PERMITIDOS:
            return False

        if estado_nuevo == "Cerrada":
            # AdminCierre y AdministradorGlobal pueden cerrar
            return rol in ("AdminCierre", "AdministradorGlobal")

        if rol == "AdministradorGlobal":
            return True
        
        if rol == "AdminCierre":
            # AdminCierre SOLO puede cerrar, no cambiar a otros estados
            return estado_nuevo == "Cerrada"

        if rol != "Responsable":
            return False

        if estado_nuevo not in ("En atención", "Solucionada"):
            return False

        bloqueadas = estado_anterior in ("Cerrada", "Archivada / No procedente")
        return not bloqueadas

    @bp.route("/denuncias/<int:did>/seguimiento", methods=["POST"])
    @login_required
    def admin_seguimiento(did):
        rol = obtener_rol()
        if rol == "Consulta":
            flash("Solo lectura.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        tipo = request.form.get("tipo_gestion")
        texto = (request.form.get("texto_seguimiento") or "").strip()
        conn = get_db()
        doc = cargar_denuncia_por_id(did)

        if not doc:
            abort(404)

        if not usuario_puede_ver_denuncia(doc, rol, session["uid"]):
            abort(403)

        if tipo not in ("observacion", "gestion"):
            flash("Tipo de gestión inválido.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        accion_txt = ""
        if tipo == "observacion":
            if len(texto) < 10:
                flash("Observación demasiado breve.", "danger")
                return redirect(url_for("admin_denuncias.admin_detalle", did=did))
            accion_txt = "OBSERVACION_INTERNA"
        elif tipo == "gestion":
            if len(texto) < 10:
                flash("Gestión demasiado breve.", "danger")
                return redirect(url_for("admin_denuncias.admin_detalle", did=did))
            accion_txt = "GESTION_REALIZADA"

        try:
            registrar_seguimiento(
                conn,
                did,
                session["uid"],
                accion=accion_txt,
                estado_anterior=doc["estado"],
                estado_nuevo=None,
                observacion=texto,
                area_id=session.get("area_id") or doc.get("area_id"),
            )

            conn.commit()
            flash("Registro de seguimiento agregado.", "success")

        except Exception:
            conn.rollback()
            flash("Error al guardar el seguimiento.", "danger")

        return redirect(url_for("admin_denuncias.admin_detalle", did=did))

    @bp.route("/denuncias/<int:did>/estado", methods=["POST"])
    @login_required
    def admin_cambia_estado(did):
        rol = obtener_rol()

        nuevo = (request.form.get("estado") or "").strip()
        texto = (request.form.get("nota_estado") or "").strip()

        if rol == "Consulta":
            flash("Sin permiso para cambiar estados.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        conn = get_db()

        doc = cargar_denuncia_por_id(did)

        if not doc:
            abort(404)

        if not usuario_puede_ver_denuncia(doc, rol, session["uid"]):
            abort(403)

        estado_ant = doc["estado"]

        if nuevo == estado_ant:
            flash("No hay cambios de estado que registrar.", "info")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        if nuevo not in ESTADOS_PERMITIDOS:
            flash("Estado inválido.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        if not _usuario_puede_cambiar_estado(rol, estado_ant, nuevo):
            if nuevo == "Cerrada" and rol != "Administrador":
                flash(
                    "Solo un usuario con perfil Administrador puede cerrar el expediente (estado «Cerrada»).",
                    "danger",
                )
            else:
                flash("Transición de estado no permitida para su perfil.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))

        fecha_cierre = doc["fecha_cierre"]

        if nuevo == "Cerrada":
            fecha_cierre = now_local()
        elif estado_ant == "Cerrada" and nuevo != "Cerrada":
            fecha_cierre = None

        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE denuncias SET estado=%s, fecha_cierre=%s, fecha_actualizacion=%s
                    WHERE id=%s
                    """,
                    (nuevo, fecha_cierre, now_local(), did),
                )
                registrar_seguimiento(
                    conn,
                    did,
                    session["uid"],
                    accion="CAMBIO_ESTADO",
                    estado_anterior=estado_ant,
                    estado_nuevo=nuevo,
                    observacion=texto if texto else None,
                    area_id=session.get("area_id") or doc.get("area_id"),
                )

            conn.commit()
            flash("Estado actualizado.", "success")
        except Exception:
            conn.rollback()
            flash("Error al actualizar el estado.", "danger")

        return redirect(url_for("admin_denuncias.admin_detalle", did=did))

    @bp.route("/usuarios")
    @login_required
    def admin_usuarios():
        # SOLO AdministradorGlobal puede gestionar usuarios
        if obtener_rol() != "AdministradorGlobal":
            flash("No tiene permiso para gestionar usuarios. Solo el Administrador Global puede hacerlo.", "danger")
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT u.id, u.nombres, u.correo, u.rol, u.activo, u.fecha_creacion, u.area_id,
                    ar.nombre AS area_nombre,
                    (
                        SELECT COUNT(*) FROM denuncias d
                        WHERE d.usuario_asignado_id = u.id
                        AND d.estado NOT IN ('Cerrada', 'Archivada / No procedente')
                    ) AS denuncias_abiertas_asignadas
                FROM usuarios u
                LEFT JOIN areas ar ON ar.id = u.area_id
                ORDER BY u.id DESC LIMIT 400
                """
            )
            u = cur.fetchall()
        candidatos_reasign = listar_usuarios_para_reasignacion_denuncias(conn)
        return render_template(
            "admin/usuarios.html",
            usuarios=u,
            roles_opts=["AdministradorGlobal", "AdminCierre", "Responsable", "Supervisor"],
            candidatos_reasign=candidatos_reasign,
        )

    @bp.route("/usuarios/crear", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_crear_usuario():
        nombre = (request.form.get("nombres") or "").strip()[:180]
        correo = (request.form.get("correo") or "").strip().lower()[:160]
        pwd = request.form.get("password") or ""
        rol = request.form.get("rol") or ""
        
        print(f"DEBUG: Intentando crear usuario - nombre={nombre}, correo={correo}, rol={rol}")
        
        conn = get_db()

        if len(nombre) < 3:
            flash("Nombre demasiado corto.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        if "@" not in correo:
            flash("Correo no válido.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        if len(pwd) < 8:
            flash("La contraseña debe tener al menos 8 caracteres.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        if rol not in ("AdministradorGlobal", "AdminCierre", "Responsable", "Supervisor"):
            flash(f"Rol inválido: {rol}", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        h = generate_password_hash(pwd)

        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO usuarios (nombres, correo, password_hash, rol, area_id, activo) VALUES (%s,%s,%s,%s,NULL,1)",
                    (nombre, correo, h, rol),
                )
                print(f"DEBUG: Usuario creado correctamente con rol {rol}")
            conn.commit()
            flash("Usuario creado correctamente.", "success")
        except Exception as e:
            conn.rollback()
            print(f"ERROR: {str(e)}")
            flash(f"No se creó el usuario: {str(e)}", "danger")

        return redirect(url_for("admin_denuncias.admin_usuarios"))

    @bp.route("/usuarios/<int:uid>/editar", methods=["GET", "POST"])
    @login_required
    def admin_editar_usuario(uid):
        # Solo AdministradorGlobal puede editar usuarios
        if obtener_rol() != "AdministradorGlobal":
            flash("No tiene permiso para editar usuarios.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))
        
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, nombres, correo, rol, area_id, activo FROM usuarios WHERE id=%s LIMIT 1",
                (uid,),
            )
            u = cur.fetchone()
        if not u:
            abort(404)

        # Obtener lista de áreas para el selector
        with conn.cursor() as cur:
            cur.execute("SELECT id, nombre FROM areas WHERE activo=1 ORDER BY nombre")
            areas = cur.fetchall()

        if request.method == "GET":
            return render_template(
                "admin/usuario_editar.html",
                u=u,
                areas=areas,
            )

        nombre = (request.form.get("nombres") or "").strip()[:180]
        correo = (request.form.get("correo") or "").strip().lower()[:160]
        pwd = (request.form.get("password") or "").strip()
        rol = request.form.get("rol") or u["rol"]
        area_id = request.form.get("area_id") or None
        if area_id == "":
            area_id = None
        else:
            try:
                area_id = int(area_id)
            except (TypeError, ValueError):
                area_id = None

        if len(nombre) < 3 or "@" not in correo:
            flash("Datos incompletos o correo inválido.", "danger")
            return redirect(url_for("admin_denuncias.admin_editar_usuario", uid=uid))

        # Validar rol
        if rol not in ("AdministradorGlobal", "AdminCierre", "Responsable", "Supervisor"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("admin_denuncias.admin_editar_usuario", uid=uid))

        # Si es Responsable o Supervisor, debe tener área asignada (opcional pero recomendado)
        if rol in ("Responsable", "Supervisor") and area_id:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM areas WHERE id=%s AND activo=1", (area_id,))
                if not cur.fetchone():
                    flash("El área seleccionada no es válida.", "danger")
                    return redirect(url_for("admin_denuncias.admin_editar_usuario", uid=uid))

        conn = get_db()
        try:
            with conn.cursor() as cur:
                if pwd:
                    if len(pwd) < 8:
                        flash("La contraseña debe tener al menos 8 caracteres.", "danger")
                        return redirect(url_for("admin_denuncias.admin_editar_usuario", uid=uid))
                    h = generate_password_hash(pwd)
                    cur.execute(
                        """
                        UPDATE usuarios SET nombres=%s, correo=%s, password_hash=%s, rol=%s, area_id=%s
                        WHERE id=%s
                        """,
                        (nombre, correo, h, rol, area_id, uid),
                    )
                else:
                    cur.execute(
                        """
                        UPDATE usuarios SET nombres=%s, correo=%s, rol=%s, area_id=%s
                        WHERE id=%s
                        """,
                        (nombre, correo, rol, area_id, uid),
                    )

            conn.commit()
            flash("Usuario actualizado correctamente.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Error al actualizar usuario: {str(e)}", "danger")

        return redirect(url_for("admin_denuncias.admin_usuarios"))

    @bp.route("/usuarios/alternar-activo", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_toggle_usuario():
        try:
            uid = int(request.form.get("usuario_id") or "")
        except ValueError:
            flash("Solicitud inválida.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        if uid == session["uid"]:
            flash("No puede desactivar su propio usuario desde aquí.", "warning")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, activo, nombres, rol, area_id FROM usuarios WHERE id=%s LIMIT 1",
                (uid,),
            )
            urow = cur.fetchone()

        if not urow:
            flash("Usuario no encontrado.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        # Reactivar: sin validación adicional
        if not urow["activo"]:
            try:
                with conn.cursor() as cur:
                    cur.execute("UPDATE usuarios SET activo=1 WHERE id=%s", (uid,))
                conn.commit()
                flash("Usuario activado.", "success")
            except Exception:
                conn.rollback()
                flash("Error al activar el usuario.", "danger")
            return redirect(url_for("admin_denuncias.admin_usuarios"))

        # Desactivar
        n_pend = contar_denuncias_abiertas_asignadas_a(conn, uid)
        rid_raw = (request.form.get("reassign_usuario_id") or "").strip()
        rid = None
        if rid_raw:
            try:
                rid = int(rid_raw)
            except ValueError:
                rid = None

        if n_pend > 0:
            if not rid:
                flash(
                    "Este usuario tiene {} expediente(s) en curso asignados. "
                    "Debe elegir a quién reasignarlos antes de desactivar la cuenta.".format(n_pend),
                    "danger",
                )
                return redirect(url_for("admin_denuncias.admin_usuarios"))
            if not candidato_valido_recibir_denuncias(conn, rid, uid):
                flash("El usuario elegido para reasignar no es válido o no puede recibir expedientes.", "danger")
                return redirect(url_for("admin_denuncias.admin_usuarios"))

        try:
            if n_pend > 0:
                reasignar_denuncias_abiertas_por_desactivacion(
                    conn, uid, rid, int(session["uid"])
                )
            with conn.cursor() as cur:
                if urow.get("rol") == "Responsable" and urow.get("area_id"):
                    cur.execute("UPDATE usuarios SET area_id=NULL WHERE id=%s", (uid,))
                cur.execute("UPDATE usuarios SET activo=0 WHERE id=%s", (uid,))
            conn.commit()
            if n_pend > 0:
                flash(
                    "Usuario desactivado. Se reasignaron {} expediente(s) en curso al usuario indicado; "
                    "quedó registrado en el historial de cada expediente.".format(n_pend),
                    "success",
                )
            else:
                flash("Usuario desactivado.", "success")
        except Exception:
            conn.rollback()
            flash("Error al desactivar o reasignar expedientes.", "danger")

        return redirect(url_for("admin_denuncias.admin_usuarios"))

    @bp.route("/categorias")
    @login_required
    def admin_categorias_panel():
        # SOLO AdministradorGlobal puede gestionar categorías
        if obtener_rol() != "AdministradorGlobal":
            flash("No tiene permiso para gestionar categorías. Solo el Administrador Global puede hacerlo.", "danger")
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        
        conn = get_db()
        areas_ls = listar_areas_activas(conn)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT u.id, u.nombres, u.area_id, a.nombre AS area_nombre
                FROM usuarios u
                LEFT JOIN areas a ON a.id = u.area_id
                WHERE u.rol = 'Responsable' AND u.activo = 1
                ORDER BY u.nombres
            """)
            responsables_ls = cur.fetchall()
        with conn.cursor() as cur:
            cur.execute("SELECT id, nombre, orden, activo FROM categorias ORDER BY orden, nombre")
            cats = cur.fetchall()
            subs = []
            if cats:
                cur.execute(
                    """
                    SELECT s.id, s.categoria_id, s.nombre, s.orden, s.activo, s.area_id_principal,
                        a.nombre AS area_pri_nombre,
                        s.usuario_id_principal,
                        ur.nombres AS usuario_pri_nombre
                    FROM subcategorias s
                    LEFT JOIN areas a ON a.id = s.area_id_principal
                    LEFT JOIN usuarios ur ON ur.id = s.usuario_id_principal
                    ORDER BY s.categoria_id, s.orden, s.nombre
                    """
                )
                subs = cur.fetchall()

        sub_dn = denuncias_por_subcategoria_map(conn)

        subs_por_cat = {}
        for s in subs:
            s["denuncias_cnt"] = sub_dn.get(int(s["id"]), 0)
            subs_por_cat.setdefault(s["categoria_id"], []).append(s)

        for c in cats:
            c["denuncias_cnt"] = contar_denuncias_para_categoria(conn, c["id"])

        return render_template(
            "admin/categorias.html",
            categorias=cats,
            subs_por_cat=subs_por_cat,
            areas_ls=areas_ls,
            responsables_ls=responsables_ls,
        )

    @bp.route("/categorias/crear", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_cat_crear():
        nom = (request.form.get("nombre") or "").strip()[:220]
        if len(nom) < 4:
            flash("Nombre de categoría demasiado corto.", "danger")
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO categorias (nombre, orden, activo) VALUES (%s, 999, 1)",
                    (nom,),
                )
            conn.commit()
            flash("Categoría creada.", "success")
        except Exception:
            conn.rollback()
            flash("No se pudo crear (¿nombre duplicado?).", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/categorias/<int:cid>/actualizar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_cat_actualizar(cid):
        nom = (request.form.get("nombre") or "").strip()[:220]
        try:
            orden = int(request.form.get("orden") or "0")
        except ValueError:
            orden = 0
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE categorias SET nombre=%s, orden=%s WHERE id=%s",
                    (nom[:220], orden, cid),
                )
            conn.commit()
            flash("Categoría actualizada.", "success")
        except Exception:
            conn.rollback()
            flash("Error al actualizar categoría.", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/categorias/<int:cid>/alternar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_cat_alternar(cid):
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE categorias SET activo = IF(activo=1,0,1) WHERE id=%s", (cid,))
            conn.commit()
            flash("Estado de categoría actualizado.", "success")
        except Exception:
            conn.rollback()
            flash("Error.", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/categorias/<int:cid>/subcategorias/crear", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_subcat_crear(cid):
        nom = (request.form.get("nombre") or "").strip()[:220]
        aprin = None
        try:
            ap = request.form.get("area_id_principal") or ""
            if ap.strip():
                aprin = int(ap)
        except ValueError:
            aprin = None
        uprin = None
        try:
            up = request.form.get("usuario_id_principal") or ""
            if up.strip():
                uprin = int(up)
        except ValueError:
            uprin = None
        if len(nom) < 4:
            flash("Nombre de subcategoría demasiado corto.", "danger")
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO subcategorias (categoria_id, nombre, orden, activo, area_id_principal, usuario_id_principal) VALUES (%s,%s,990,1,%s,%s)",
                    (cid, nom[:220], aprin, uprin),
                )
            conn.commit()
            flash("Subcategoría creada.", "success")
        except Exception:
            conn.rollback()
            flash("No se pudo crear subcategoría (¿nombre duplicado?).", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/subcategorias/<int:sid>/actualizar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_subcat_actualizar(sid):
        nom = (request.form.get("nombre") or "").strip()[:220]
        try:
            orden = int(request.form.get("orden") or "0")
        except ValueError:
            orden = 0
        aprin = None
        try:
            ap = request.form.get("area_id_principal") or ""
            if ap.strip():
                aprin = int(ap)
        except ValueError:
            aprin = None
        uprin = None
        try:
            up = request.form.get("usuario_id_principal") or ""
            if up.strip():
                uprin = int(up)
        except ValueError:
            uprin = None
        cid = None
        try:
            cid = int(request.form.get("categoria_id") or "")
        except ValueError:
            pass
        conn = get_db()
        try:
            with conn.cursor() as cur:
                if cid:
                    cur.execute(
                        "UPDATE subcategorias SET nombre=%s, orden=%s, area_id_principal=%s, usuario_id_principal=%s, categoria_id=%s WHERE id=%s",
                        (nom[:220], orden, aprin, uprin, cid, sid),
                    )
                else:
                    cur.execute(
                        "UPDATE subcategorias SET nombre=%s, orden=%s, area_id_principal=%s, usuario_id_principal=%s WHERE id=%s",
                        (nom[:220], orden, aprin, uprin, sid),
                    )
            conn.commit()
            flash("Subcategoría actualizada.", "success")
        except Exception:
            conn.rollback()
            flash("Error al actualizar subcategoría.", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/subcategorias/<int:sid>/alternar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_subcat_alternar(sid):
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE subcategorias SET activo = IF(activo=1,0,1) WHERE id=%s", (sid,))
            conn.commit()
            flash("Subcategoría actualizada.", "success")
        except Exception:
            conn.rollback()
            flash("Error.", "danger")
        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/categorias/<int:cid>/eliminar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_cat_eliminar(cid):
        conn = get_db()
        n_exp = contar_denuncias_para_categoria(conn, cid)
        if n_exp > 0:
            flash(
                "No puede eliminar la categoría: hay {} expediente(s) de denuncia o queja vinculado(s)."
                "".format(n_exp),
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))

        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM categorias WHERE id=%s", (cid,))
                if cur.rowcount == 0:
                    flash("No se encontró la categoría.", "warning")
                else:
                    flash("Categoría y sus subcategorías eliminadas del catálogo.", "success")
            conn.commit()
        except Exception:
            conn.rollback()
            flash("No se pudo eliminar la categoría (restricción en base de datos).", "danger")

        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/subcategorias/<int:sid>/eliminar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_subcat_eliminar(sid):
        try:
            cid_check = int(request.form.get("categoria_id") or "")
        except ValueError:
            flash("Datos incompletos para eliminar.", "danger")
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))

        conn = get_db()
        cat_real = None
        with conn.cursor() as cur:
            cur.execute("SELECT categoria_id FROM subcategorias WHERE id=%s LIMIT 1", (sid,))
            row = cur.fetchone()
            if not row:
                flash("No se encontró la subcategoría.", "warning")
                return redirect(url_for("admin_denuncias.admin_categorias_panel"))
            cat_real = row["categoria_id"]

        if int(cat_real) != int(cid_check):
            flash("Solicitud no válida (categoría no coincide).", "danger")
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))

        n_exp = contar_denuncias_para_subcategoria(conn, sid)
        if n_exp > 0:
            flash(
                "No puede eliminar esta subcategoría: hay {} expediente(s) vinculado(s).".format(n_exp),
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_categorias_panel"))

        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM subcategorias WHERE id=%s", (sid,))
                if cur.rowcount == 0:
                    flash("No se encontró la subcategoría.", "warning")
                else:
                    flash("Subcategoría eliminada del catálogo.", "success")
            conn.commit()
        except Exception:
            conn.rollback()
            flash("No se pudo eliminar (restricción en la base de datos).", "danger")

        return redirect(url_for("admin_denuncias.admin_categorias_panel"))

    @bp.route("/areas")
    @login_required
    def admin_areas_panel():
        # SOLO AdministradorGlobal puede gestionar áreas
        if obtener_rol() != "AdministradorGlobal":
            flash("No tiene permiso para gestionar áreas. Solo el Administrador Global puede hacerlo.", "danger")
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        
        conn = get_db()
        dmap = denuncias_por_area_map(conn)
        responsables_catalog = listar_responsables_catalogo(conn)
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT a.*,
                    r.id AS responsable_uid,
                    r.nombres AS responsable_nombres,
                    r.correo AS responsable_correo
                FROM areas a
                LEFT JOIN usuarios r ON r.area_id = a.id AND r.rol='Responsable' AND r.activo=1
                ORDER BY a.nombre
                """
            )
            ars = cur.fetchall()
        for a in ars:
            a["denuncias_cnt"] = dmap.get(int(a["id"]), 0)
        return render_template(
            "admin/areas.html",
            areas=ars,
            responsables_catalog=responsables_catalog,
        )

    @bp.route("/areas/responsable-rapido", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_area_responsable_rapido():
        """Alta rápida de Responsable sin área. Respuesta JSON para el modal de áreas."""
        data = request.get_json(silent=True) or {}
        nombre = (data.get("nombres") or "").strip()[:180]
        correo = (data.get("correo") or "").strip().lower()[:160]
        pwd = data.get("password") or ""

        if len(nombre) < 3:
            return jsonify(ok=False, error="El nombre es obligatorio (mínimo 3 caracteres)."), 400
        if "@" not in correo:
            return jsonify(ok=False, error="Correo no válido."), 400
        if len(pwd) < 8:
            return jsonify(ok=False, error="La contraseña debe tener al menos 8 caracteres."), 400

        conn = get_db()
        h = generate_password_hash(pwd)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO usuarios (nombres, correo, password_hash, rol, area_id, activo)
                    VALUES (%s,%s,%s,'Responsable',NULL,1)
                    """,
                    (nombre, correo, h),
                )
                new_id = cur.lastrowid
            conn.commit()
            return jsonify(
                ok=True,
                id=int(new_id),
                nombres=nombre,
                correo=correo,
                area_id=None,
            )
        except IntegrityError:
            conn.rollback()
            return jsonify(ok=False, error="Ese correo ya está registrado."), 409
        except Exception:
            conn.rollback()
            return jsonify(ok=False, error="No se pudo crear el usuario."), 500

    @bp.route("/areas/crear", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_area_crear():
        nom = (request.form.get("nombre") or "").strip()[:220]
        rid = request.form.get("responsable_usuario_id") or ""
        if len(nom) < 4:
            flash("Indique un nombre de área válido (mínimo 4 caracteres).", "danger")
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        try:
            rid_int = int(rid)
        except (TypeError, ValueError):
            flash("Debe elegir un responsable para el área.", "danger")
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        conn = get_db()
        if not usuario_responsable_disponible_para_area(conn, rid_int, None):
            flash(
                "El responsable elegido ya está asignado a otra área. Use solo usuarios disponibles o cree uno nuevo.",
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO areas (nombre, activo) VALUES (%s, 1)",
                    (nom[:220],),
                )
                new_id = cur.lastrowid
            ok, err = asignar_responsable_a_area(conn, new_id, rid_int)
            if not ok:
                conn.rollback()
                flash(err or "No se pudo vincular el responsable.", "danger")
                return redirect(url_for("admin_denuncias.admin_areas_panel"))
            conn.commit()
            flash("Área institucional creada.", "success")
        except IntegrityError as e:
            conn.rollback()
            code = e.args[0] if e.args else 0
            if int(code) == 1062:
                flash("Ya existe un área con ese nombre. Elija otro nombre.", "danger")
            else:
                flash("No se pudo crear el área (restricción en la base de datos).", "danger")
        except Exception as e:
            conn.rollback()
            msg = str(e.args[1]) if len(e.args) > 1 else str(e)
            if "contacto" in msg.lower() or "Unknown column" in msg:
                flash(
                    "La base de datos aún tiene columnas antiguas en «areas». Ejecute el script "
                    "migrate_areas_drop_contacto.sql (o cree la base desde schema.sql actualizado).",
                    "danger",
                )
            else:
                flash("Error al crear el área: {}.".format(msg[:200]), "danger")
        return redirect(url_for("admin_denuncias.admin_areas_panel"))

    @bp.route("/areas/<int:aid>/actualizar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_area_actualizar(aid):
        nom = (request.form.get("nombre") or "").strip()[:220]
        rid = request.form.get("responsable_usuario_id") or ""
        if len(nom) < 4:
            flash("Indique un nombre de área válido (mínimo 4 caracteres).", "danger")
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        try:
            rid_int = int(rid)
        except (TypeError, ValueError):
            flash("Debe elegir un responsable para el área.", "danger")
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        conn = get_db()
        if not usuario_responsable_disponible_para_area(conn, rid_int, aid):
            flash(
                "El responsable elegido ya está asignado a otra área distinta. Elija un usuario disponible.",
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_areas_panel"))
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE areas SET nombre=%s, fecha_actualizacion=%s WHERE id=%s",
                    (nom[:220], now_local(), aid),
                )
            ok, err = asignar_responsable_a_area(conn, aid, rid_int)
            if not ok:
                conn.rollback()
                flash(err or "No se pudo actualizar el responsable del área.", "danger")
                return redirect(url_for("admin_denuncias.admin_areas_panel"))
            conn.commit()
            flash("Área actualizada.", "success")
        except Exception:
            conn.rollback()
            flash("Error al actualizar área.", "danger")
        return redirect(url_for("admin_denuncias.admin_areas_panel"))

    @bp.route("/areas/<int:aid>/alternar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_area_alternar(aid):
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE areas SET activo = IF(activo=1,0,1) WHERE id=%s", (aid,))
            conn.commit()
            flash("Estado del área actualizado.", "success")
        except Exception:
            conn.rollback()
            flash("Error.", "danger")
        return redirect(url_for("admin_denuncias.admin_areas_panel"))

    @bp.route("/areas/<int:aid>/eliminar", methods=["POST"])
    @login_required
    @roles_required("Administrador")
    def admin_area_eliminar(aid):
        conn = get_db()
        n_exp = contar_denuncias_para_area(conn, aid)
        if n_exp > 0:
            flash(
                "No puede eliminar esta área: hay {} expediente(s) asignados a ese área."
                "".format(n_exp),
                "danger",
            )
            return redirect(url_for("admin_denuncias.admin_areas_panel"))

        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM areas WHERE id=%s", (aid,))
                if cur.rowcount == 0:
                    flash("No se encontró el área.", "warning")
                else:
                    flash("Área institucional eliminada.", "success")
            conn.commit()
        except Exception:
            conn.rollback()
            flash("No se pudo eliminar el área (restricción en base de datos).", "danger")

        return redirect(url_for("admin_denuncias.admin_areas_panel"))
    
    @bp.route("/denuncias/<int:did>/toggle-activo", methods=["POST"])
    @login_required
    def admin_toggle_denuncia_activo(did):
        """Solo Administrador puede desactivar/reactivar denuncias con comentario obligatorio."""
        rol = obtener_rol()
        
        # SOLO ADMINISTRADOR
        if rol != "Administrador":
            flash("No tiene permiso para realizar esta acción.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        
        conn = get_db()
        doc = cargar_denuncia_por_id(did)
        if not doc:
            abort(404)
        
        # Obtener comentario (obligatorio para desactivar)
        comentario = (request.form.get("comentario_desactivacion") or "").strip()
        
        # Determinar la acción
        nuevo_estado = 0 if doc.get("activo") == 1 else 1
        
        # Si va a desactivar (activo=0), el comentario es obligatorio
        if nuevo_estado == 0 and len(comentario) < 10:
            flash("Para desactivar una denuncia debe indicar el motivo (mínimo 10 caracteres).", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE denuncias SET activo = %s, fecha_actualizacion = %s WHERE id = %s",
                    (nuevo_estado, now_local(), did)
                )
            
            accion = "DESACTIVACION" if nuevo_estado == 0 else "REACTIVACION"
            observacion = f"Denuncia {accion.lower()} por Administrador {session.get('nombres')}. Motivo: {comentario}"
            
            registrar_seguimiento(
                conn,
                did,
                session["uid"],
                accion=accion,
                estado_anterior=doc["estado"],
                estado_nuevo=doc["estado"],
                observacion=observacion,
                area_id=doc.get("area_id"),
            )
            conn.commit()
            
            if nuevo_estado == 0:
                flash(f"Denuncia {doc['codigo']} desactivada correctamente.", "success")
            else:
                flash(f"Denuncia {doc['codigo']} reactivada correctamente.", "success")
                
        except Exception as e:
            conn.rollback()
            flash("Error al cambiar el estado de la denuncia.", "danger")
        
        return redirect(url_for("admin_denuncias.admin_detalle", did=did))
    
    @bp.route("/denuncias/<int:did>/recuperar-reasignacion", methods=["POST"])
    @login_required
    def admin_recuperar_reasignacion(did):
        """Permite volver al responsable anterior si el responsable actual se equivocó al reasignar."""
        rol = obtener_rol()
        
        # Solo Responsable puede recuperar reasignaciones
        if rol != "Responsable":
            flash("No tiene permiso para realizar esta acción.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        
        conn = get_db()
        doc = cargar_denuncia_por_id(did)
        if not doc:
            abort(404)
        
        if not usuario_puede_ver_denuncia(doc, rol, session["uid"]):
            abort(403)
        
        responsable_anterior = doc.get("responsable_anterior_id")
        
        if not responsable_anterior:
            flash("No hay un responsable anterior para recuperar.", "warning")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        
        # Obtener información del responsable anterior
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, nombres, correo, area_id FROM usuarios WHERE id=%s AND activo=1",
                (responsable_anterior,)
            )
            anterior = cur.fetchone()
        
        if not anterior:
            flash("El responsable anterior ya no está activo. No se puede recuperar.", "danger")
            return redirect(url_for("admin_denuncias.admin_detalle", did=did))
        
        comentario = request.form.get("comentario_recuperacion") or "Recuperación de reasignación anterior"
        estado_ant = doc["estado"]
        
        try:
            # Actualizar denuncia al responsable anterior
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE denuncias
                    SET usuario_asignado_id=%s, area_id=%s, fecha_actualizacion=%s, responsable_anterior_id=NULL
                    WHERE id=%s
                    """,
                    (anterior["id"], anterior["area_id"], now_local(), did),
                )
            
            registrar_seguimiento(
                conn,
                did,
                session["uid"],
                accion="RECUPERACION_REASIGNACION",
                estado_anterior=estado_ant,
                estado_nuevo=estado_ant,
                observacion=f"Se recuperó la reasignación. Responsable restaurado: {anterior['nombres']}. Motivo: {comentario}",
                area_id=anterior["area_id"],
            )
            conn.commit()
            
            flash(f"✅ Se ha recuperado la reasignación. El expediente vuelve a estar a cargo de {anterior['nombres']}.", "success")
            
        except Exception as e:
            conn.rollback()
            flash("Error al recuperar la reasignación.", "danger")
        
        return redirect(url_for("admin_denuncias.admin_detalle", did=did))

    @bp.route("/denuncias/<int:did>/recuperar-mis-reasignacion", methods=["POST"])
    @login_required
    def admin_recuperar_mis_reasignacion(did):
        """Permite al responsable recuperar una denuncia que él mismo reasignó por error."""
        rol = obtener_rol()
        
        if rol != "Responsable":
            flash("No tiene permiso para realizar esta acción.", "danger")
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        
        conn = get_db()
        doc = cargar_denuncia_por_id(did)
        if not doc:
            abort(404)
        
        # Verificar que el usuario actual fue quien reasignó esta denuncia
        if doc.get("reasignado_por_id") != session["uid"]:
            flash("Solo puede recuperar denuncias que usted mismo reasignó.", "danger")
            return redirect(url_for("admin_denuncias.admin_mis_reasignaciones"))
        
        # Obtener el responsable anterior
        responsable_anterior = doc.get("responsable_anterior_id")
        if not responsable_anterior:
            flash("No hay un responsable anterior para recuperar.", "warning")
            return redirect(url_for("admin_denuncias.admin_mis_reasignaciones"))
        
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, nombres, area_id FROM usuarios WHERE id=%s AND activo=1",
                (responsable_anterior,)
            )
            anterior = cur.fetchone()
        
        if not anterior:
            flash("El responsable anterior ya no está activo. No se puede recuperar.", "danger")
            return redirect(url_for("admin_denuncias.admin_mis_reasignaciones"))
        
        comentario = request.form.get("comentario_recuperacion") or "Recuperación por error en reasignación"
        
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE denuncias
                    SET usuario_asignado_id=%s, area_id=%s, fecha_actualizacion=%s, 
                        responsable_anterior_id=NULL, reasignado_por_id=NULL, fecha_reasignacion=NULL
                    WHERE id=%s
                    """,
                    (anterior["id"], anterior["area_id"], now_local(), did),
                )
            
            registrar_seguimiento(
                conn,
                did,
                session["uid"],
                accion="RECUPERACION_MIS_REASIGNACION",
                estado_anterior=doc["estado"],
                estado_nuevo=doc["estado"],
                observacion=f"El responsable recuperó la denuncia que había reasignado. Motivo: {comentario}",
                area_id=anterior["area_id"],
            )
            conn.commit()
            
            flash(f"✅ Ha recuperado la denuncia. Ahora está nuevamente a su cargo.", "success")
            
        except Exception as e:
            conn.rollback()
            flash("Error al recuperar la denuncia.", "danger")
        
        return redirect(url_for("admin_denuncias.admin_mis_reasignaciones"))
    
    @bp.route("/mis-reasignaciones")
    @login_required
    def admin_mis_reasignaciones():
        """Bandeja de denuncias que el responsable reasignó recientemente (últimas 48 horas)."""
        rol = obtener_rol()
        
        if rol != "Responsable":
            flash("No tiene permiso para acceder.", "danger")
            return redirect(url_for("admin_denuncias.admin_dashboard"))
        
        conn = get_db()
        uid = session["uid"]
        
        # Denuncias que este responsable reasignó en las últimas 48 horas
        sql = """
            SELECT d.id, d.codigo, d.categoria, d.subcategoria, d.estado, d.prioridad,
                d.fecha_reasignacion, u.nombres AS reasignado_a,
                d.fecha_creacion
            FROM denuncias d
            LEFT JOIN usuarios u ON u.id = d.usuario_asignado_id
            WHERE d.reasignado_por_id = %s 
            AND d.fecha_reasignacion > DATE_SUB(NOW(), INTERVAL 48 HOUR)
            AND d.responsable_anterior_id IS NOT NULL
            ORDER BY d.fecha_reasignacion DESC
        """
        
        with conn.cursor() as cur:
            cur.execute(sql, (uid,))
            reasignaciones = cur.fetchall()
        
        return render_template(
            "admin/mis_reasignaciones.html",
            reasignaciones=reasignaciones,
            tiene_registros=len(reasignaciones) > 0
        )

    @bp.route("/reportes")
    @login_required
    def admin_reportes():
        rol = session["rol"]
        uid = session["uid"]

        filt, args = listar_where_base(rol, uid)

        conn = get_db()
        totals = []

        sql = """
        SELECT estado, COUNT(*) AS c
        FROM denuncias d
        WHERE 1=1 {filt}
        GROUP BY estado
        ORDER BY c DESC
        """.format(
            filt=filt
        )

        with conn.cursor() as cur:
            cur.execute(sql, tuple(args))
            totals = cur.fetchall()

        datos_por_area = None
        total_area = 0
        datos_por_categoria = None
        total_categoria = 0
        datos_por_subcat = None
        total_subcat = 0
        if rol in ("AdministradorGlobal", "AdminCierre"):
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT a.nombre, COUNT(d.id) AS c
                    FROM areas a
                    LEFT JOIN denuncias d ON d.area_id = a.id AND d.activo = 1
                    GROUP BY a.id, a.nombre
                    ORDER BY c DESC, a.nombre
                    """
                )
                datos_por_area = cur.fetchall()
            for row in datos_por_area:
                row["c"] = int(row["c"])
            total_area = sum(r["c"] for r in datos_por_area)
            for row in datos_por_area:
                row["pct"] = round(row["c"] / total_area * 100, 1) if total_area > 0 else 0.0

            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT cat.nombre, COUNT(d.id) AS c
                    FROM categorias cat
                    LEFT JOIN denuncias d ON d.categoria_id = cat.id AND d.activo = 1
                    GROUP BY cat.id, cat.nombre
                    ORDER BY c DESC, cat.nombre
                    """
                )
                datos_por_categoria = cur.fetchall()
            for row in datos_por_categoria:
                row["c"] = int(row["c"])
            total_categoria = sum(r["c"] for r in datos_por_categoria)
            for row in datos_por_categoria:
                row["pct"] = round(row["c"] / total_categoria * 100, 1) if total_categoria > 0 else 0.0

            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT sub.nombre, COUNT(d.id) AS c
                    FROM subcategorias sub
                    LEFT JOIN denuncias d ON d.subcategoria_id = sub.id AND d.activo = 1
                    GROUP BY sub.id, sub.nombre
                    ORDER BY c DESC, sub.nombre
                    """
                )
                datos_por_subcat = cur.fetchall()
            for row in datos_por_subcat:
                row["c"] = int(row["c"])
            total_subcat = sum(r["c"] for r in datos_por_subcat)
            for row in datos_por_subcat:
                row["pct"] = round(row["c"] / total_subcat * 100, 1) if total_subcat > 0 else 0.0

        return render_template(
            "admin/reportes.html",
            tabla=totals,
            datos_por_area=datos_por_area,
            total_area=total_area,
            datos_por_categoria=datos_por_categoria,
            total_categoria=total_categoria,
            datos_por_subcat=datos_por_subcat,
            total_subcat=total_subcat,
        )

    @bp.route("/reportes/excel")
    @login_required
    def excel_report():
        rol = session["rol"]
        uid = session["uid"]

        filt, args = listar_where_base(rol, uid)

        sql = """
        SELECT d.codigo, d.categoria, d.subcategoria, d.prioridad, d.estado,
               COALESCE(u.nombres,'') AS responsable,
               COALESCE(ar.nombre,'') AS area_institucional,
               d.fecha_creacion, d.fecha_cierre
        FROM denuncias d
        LEFT JOIN usuarios u ON u.id=d.usuario_asignado_id
        LEFT JOIN areas ar ON ar.id=d.area_id
        WHERE 1=1 {filt}
        ORDER BY d.fecha_creacion DESC
        """.format(
            filt=filt
        )

        conn = get_db()
        rows = []

        with conn.cursor() as cur:
            cur.execute(sql, tuple(args))
            rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Denuncias"

        hdr = [
            "código",
            "categoría",
            "subcategoría",
            "prioridad",
            "estado",
            "responsable cuenta",
            "área institucional",
            "fecha creación",
            "fecha cierre",
        ]

        for col, nm in enumerate(hdr, start=1):
            ws.cell(row=1, column=col, value=nm)

        ridx = 2

        def as_date(v):
            if v is None:
                return ""

            try:
                if isinstance(v, str):
                    return v[:19]
                if hasattr(v, "strftime"):
                    return v.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(v)
            return ""

        for rec in rows:
            ws.cell(row=ridx, column=1, value=str(rec["codigo"]))
            ws.cell(row=ridx, column=2, value=str(rec["categoria"]))
            ws.cell(row=ridx, column=3, value=str(rec["subcategoria"]))
            ws.cell(row=ridx, column=4, value=str(rec["prioridad"]))
            ws.cell(row=ridx, column=5, value=str(rec["estado"]))
            ws.cell(row=ridx, column=6, value=str(rec["responsable"]))
            ws.cell(row=ridx, column=7, value=str(rec.get("area_institucional") or ""))
            ws.cell(row=ridx, column=8, value=as_date(rec["fecha_creacion"]))
            ws.cell(row=ridx, column=9, value=as_date(rec["fecha_cierre"]))
            ridx += 1

        for col in range(1, len(hdr) + 1):
            col_letter = get_column_letter(col)
            max_len = 10
            for row in ws.iter_rows(min_row=1, max_row=min(ridx, 500)):
                cel = row[col - 1]
                vl = getattr(cel, "value", None)
                ln = len(str(vl)) if vl is not None else 0
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = "reporte_denuncias_mmq_{}.xlsx".format(
            datetime.now(get_tz()).strftime("%Y%m%d_%H%M")
        )

        rp = Response(
            stream.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        rp.headers["Content-Disposition"] = 'attachment; filename="{}"'.format(filename.replace('"', ""))
        return rp


app = create_app()


@app.route("/")
def raiz_redirect():
    return redirect(url_for("public_denuncias.index"))


if __name__ == "__main__":
    app.run(debug=__import__("config").FLASK_DEBUG)